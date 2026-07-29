"""Service for Excel file processing."""

import logging
from typing import Optional

from openpyxl import load_workbook

from .models import Project, State, County, City, UploadedFile
from .exceptions import ExcelProcessingError
from .page_service import PageService

logger = logging.getLogger(__name__)


class ExcelProcessingService:
    """Service for processing uploaded Excel files."""

    def __init__(self, page_service: Optional[PageService] = None) -> None:
        self._page_service = page_service or PageService()

    def process_uploaded_file(
        self, uploaded_file: UploadedFile, prompt: str
    ) -> None:
        """Process an uploaded Excel file to create city pages.

        Args:
            uploaded_file: The UploadedFile instance.
            prompt: Additional prompt guidance.
        """
        workbook = load_workbook(uploaded_file.file.path)
        sheet = workbook.active

        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index == 0 or not row[0] or not row[3] or not row[5]:
                continue

            state, _ = State.objects.get_or_create(
                name=row[3], abbreviation=row[2], project=uploaded_file.project
            )
            county, _ = County.objects.get_or_create(
                name=row[5], state=state
            )
            city, _ = City.objects.get_or_create(name=row[0], county=county)

            city.status = CityStatus.QUEUED.value
            city.save()

            logger.info('Processing city: %s', city.name)
            self._page_service.create_city_page(
                uploaded_file.project,
                city,
                uploaded_file.temperature or 0.0,
                prompt,
            )